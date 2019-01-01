#!/usr/bin/env python3
import openpyxl
import os
import time
from selenium import webdriver
from collections import OrderedDict

# Open a web browser and navigate to URL
browser = webdriver.Chrome()
browser.get('https://myhealth.alberta.ca/Alberta/Pages/Heart-Disease-Risk-Calculator.aspx?fbclid=IwAR1e_tA9dNsH5SkT8yg3V0_X7Whn9ZWbDqBRI4nHwHcoS5Nc_7ZFBXPscSc')

# Create and setup Excel Workbook
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 'Age'
sheet['B1'] = 'Gender'
sheet['C1'] = 'Smoker'
sheet['D1'] = 'Diabetic'
sheet['E1'] = 'Total Cholesterol'
sheet['F1'] = 'HDL'
sheet['G1'] = 'Systolic BP'
sheet['H1'] = 'BP Med'
sheet['I1'] = 'Family History'
sheet['J1'] = 'Cardiovascular Risk'
sheet['K1'] = 'Heart Age'

# Create Dictionaries of CSS Selectors
ageElem = browser.find_element_by_css_selector('#mhaHdrcAge-1')
genderSelList = OrderedDict({'#mhaGenderMale':'Male','#mhaGenderFemale':'Female'})
smokerSelList = OrderedDict({'#mhaSmokerYes':'Yes','#mhaSmokerNo':'No'})
diabetesSelList = OrderedDict({'#mhaDiabeticYes':'Yes','#mhaDiabeticNo':'No'})
totCholSelList = OrderedDict({'#hrdcCholesterolLow':'Low','#hrdcCholesterolNormal':'Normal','#hrdcCholesterolHigh':'High','#hrdcCholesterolNil':'Don\'t Know'})
hdlcSelList = OrderedDict({'#hdrcTotalHdlcLow':'Low','#hdrcTotalHdlcNormal':'Normal','#hdrcTotalHdlcHigh':'High','#hdrcTotalHdlcNil':'Don\'t Know'})
sysBPSelList = OrderedDict({'#hdrcSystolicLow':'Low','#hdrcSystolicNormal':'Normal','#hdrcSystolicHigh':'High','#hdrcSystolicNil':'Don\'t Know'})
bpMedSelList = OrderedDict({'#hdrcMedicationYes':'Yes','#hdrcMedicationNo':'No'})
cardDisSelList = OrderedDict({'#hdrcDiagnosisYes':'Yes','#hdrcDiagnosisNo':'No'})

# Set up Navigation Buttons
pgForward = browser.find_element_by_css_selector('#mhaRiskform > div.actions.clearfix.bottom > ul > li.last-child > a > span.icon > i')
pgBackward = browser.find_element_by_css_selector('#mhaRiskform > div.actions.clearfix.bottom > ul > li:nth-child(1) > a > span.icon > i')

# Save entries
entry = 2

# Main Loop 
for iAge in range(30,31):
    for iGender in genderSelList.keys():
        for iSmoker in smokerSelList.keys():
            for iDiabetes in diabetesSelList.keys():
                for iTotChol in totCholSelList.keys():
                    for iHDLC in hdlcSelList.keys():
                        for iSysBP in sysBPSelList.keys():
                            for iBPMed in bpMedSelList.keys():
                                for iCardDis in cardDisSelList.keys():
                                    # Select appropriate Form answers
                                    time.sleep(1)
                                    ageElem.clear()
                                    ageElem.send_keys(str(iAge))
                                    genderElem = browser.find_element_by_css_selector(iGender)
                                    genderElem.click()
                                    smokerElem = browser.find_element_by_css_selector(iSmoker)
                                    smokerElem.click()
                                    diabetesElem = browser.find_element_by_css_selector(iDiabetes)
                                    diabetesElem.click()
                                    pgForward.click()
                                    time.sleep(1)
                                    totCholElem = browser.find_element_by_css_selector(iTotChol)
                                    totCholElem.click()
                                    hdlcElem = browser.find_element_by_css_selector(iHDLC)
                                    hdlcElem.click()
                                    pgForward.click()
                                    time.sleep(1)
                                    sysBPElem = browser.find_element_by_css_selector(iSysBP)
                                    sysBPElem.click()
                                    bpMedElem = browser.find_element_by_css_selector(iBPMed)
                                    bpMedElem.click()
                                    pgForward.click()
                                    time.sleep(1)
                                    cardDisElem = browser.find_element_by_css_selector(iCardDis)
                                    cardDisElem.click()
                                    pgForward.click()
                                    time.sleep(2)

                                    # Save results in an Excel Workbook
                                    sheet.cell(row=entry, column=1).value = iAge
                                    sheet.cell(row=entry, column=2).value = genderSelList[iGender]
                                    sheet.cell(row=entry, column=3).value = smokerSelList[iSmoker]
                                    sheet.cell(row=entry, column=4).value = diabetesSelList[iDiabetes]
                                    sheet.cell(row=entry, column=5).value = totCholSelList[iTotChol]
                                    sheet.cell(row=entry, column=6).value = hdlcSelList[iHDLC]
                                    sheet.cell(row=entry, column=7).value = sysBPSelList[iSysBP]
                                    sheet.cell(row=entry, column=8).value = bpMedSelList[iBPMed]
                                    sheet.cell(row=entry, column=9).value = cardDisSelList[iCardDis]
                                    sheet.cell(row=entry, column=10).value = browser.find_element_by_css_selector('#mhaRiskform-p-4 > div > div.mha-hdrc-results > div:nth-child(1) > ul > li:nth-child(1) > span').text
                                    sheet.cell(row=entry, column=11).value = browser.find_element_by_css_selector('#mhaRiskform-p-4 > div > div.mha-hdrc-results > div:nth-child(1) > ul > li:nth-child(2) > span').text

                                    # Increment 
                                    entry += 1

                                    # Return to beginning of Form
                                    for i in range (4):
                                        pgBackward.click()
                                        time.sleep(1)
                                        
# Save Workbook
os.chdir('C:\\Users\\Owner\\Desktop')
wb.save('hdr_calc.xlsx')
                        

                
                
                

        
    
